#!/usr/bin/env python3
"""
Report Export Module
====================
Exports analysis results and KPIs to Excel, CSV, and summary reports.
Generates multi-sheet Excel workbooks with formatted data.
"""

import pandas as pd
import numpy as np
import os
import sys
from datetime import datetime
from typing import Optional, Dict, Any, List

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from analytics_utils import (
    load_raw_data, load_cleaned_data, standardize_columns,
    calculate_kpis, print_section,
    DATA_RAW_DIR, DATA_CLEANED_DIR, DATA_PROCESSED_DIR,
    REPORTS_EXCEL_DIR, PROJECT_ROOT
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers
    )
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Excel export disabled.")
    print("Install with: pip install openpyxl")


# Styling constants for Excel
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1E293B")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="2563EB")
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB")
)
ALT_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")


def apply_header_style(ws, row, max_col):
    """Apply styling to header row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_data_styles(ws, start_row, end_row, max_col):
    """Apply styling to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (row - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def export_kpi_sheet(wb: Workbook, kpis: Dict[str, Any]):
    """Create KPI summary sheet."""
    ws = wb.active
    ws.title = "KPI Summary"

    ws.cell(row=1, column=1, value="HR Analytics - KPI Summary").font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws.cell(row=2, column=1, value=f"Generated: {kpis.get('calc_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}").font = Font(italic=True, color="64748B")

    # Write KPIs
    row = 4
    ws.cell(row=row, column=1, value="Metric").font = SECTION_FONT
    ws.cell(row=row, column=2, value="Value").font = SECTION_FONT

    kpi_groups = {
        "Workforce Overview": ["total_employees", "active_employees", "attrited_employees",
                                "department_count", "male_count", "female_count"],
        "Attrition": ["attrition_rate", "overtime_percentage"],
        "Compensation": ["avg_salary", "median_salary", "min_salary", "max_salary",
                         "total_salary_budget"],
        "Demographics": ["avg_age", "avg_tenure", "female_percentage"],
        "Satisfaction & Performance": ["avg_job_satisfaction", "avg_work_life_balance",
                                       "avg_performance_rating", "avg_training_hours"],
        "Promotion": ["promotion_rate", "avg_years_since_promotion"]
    }

    row = 5
    for group_name, metric_keys in kpi_groups.items():
        ws.cell(row=row, column=1, value=group_name).font = SECTION_FONT
        row += 1
        for key in metric_keys:
            if key in kpis:
                ws.cell(row=row, column=1, value=key.replace("_", " ").title())
                val = kpis[key]
                if isinstance(val, float):
                    ws.cell(row=row, column=2, value=round(val, 2))
                else:
                    ws.cell(row=row, column=2, value=val)
                ws.cell(row=row, column=2).number_format = '#,##0.00' if isinstance(val, float) else '@'
                row += 1
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def export_department_sheet(wb: Workbook, df: pd.DataFrame):
    """Create department analysis sheet."""
    ws = wb.create_sheet("Department Analysis")

    ws.cell(row=1, column=1, value="Department Analysis").font = TITLE_FONT
    ws.merge_cells("A1:G1")

    # Pivot table: department summary
    dept_summary = df.groupby("department", observed=True).agg(
        Count=("employee_id", "count"),
        Avg_Salary=("monthly_income", "mean"),
        Avg_Age=("age", "mean"),
        Avg_Tenure=("years_at_company", "mean"),
        Attrition_Rate=("attrition", lambda x: (x == "Yes").mean() * 100),
        Avg_Performance=("performance_rating", "mean")
    ).round(2).reset_index()
    dept_summary.columns = ["Department", "Count", "Avg Salary", "Avg Age",
                            "Avg Tenure", "Attrition Rate %", "Avg Performance"]

    for r_idx, row in enumerate(dataframe_to_rows(dept_summary, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, 3, len(dept_summary.columns))
    apply_data_styles(ws, 4, 3 + len(dept_summary), len(dept_summary.columns))

    # Adjust column widths safely (skip merged cells)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except AttributeError:
                pass
        if max_length > 0:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_length + 3, 25)


def export_attrition_sheet(wb: Workbook, df: pd.DataFrame):
    """Create attrition analysis sheet."""
    ws = wb.create_sheet("Attrition Analysis")

    ws.cell(row=1, column=1, value="Attrition Analysis").font = TITLE_FONT
    ws.merge_cells("A1:G1")

    # Attrition by department
    dept_attr = df[df["attrition"] == "Yes"].groupby("department", observed=True).agg(
        Count=("employee_id", "count"),
        Avg_Salary=("monthly_income", "mean"),
        Avg_Age=("age", "mean"),
        Avg_Tenure=("years_at_company", "mean"),
        Avg_Satisfaction=("job_satisfaction", "mean")
    ).round(2).reset_index()
    dept_attr.columns = ["Department", "Attrited Count", "Avg Salary",
                         "Avg Age", "Avg Tenure", "Avg Satisfaction"]

    ws.cell(row=3, column=1, value="Attrition by Department").font = SECTION_FONT
    for r_idx, row in enumerate(dataframe_to_rows(dept_attr, index=False, header=True), 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, 4, len(dept_attr.columns))
    apply_data_styles(ws, 5, 4 + len(dept_attr), len(dept_attr.columns))

    # Attrition by factors
    factors = {"Gender": "gender", "Job Level": "job_level",
               "Marital Status": "marital_status", "Overtime": "overtime"}
    row_start = 6 + len(dept_attr)

    for label, col in factors.items():
        ws.cell(row=row_start, column=1, value=f"Attrition by {label}").font = SECTION_FONT
        factor_data = df.groupby(col, observed=True)["attrition"].apply(
            lambda x: (x == "Yes").mean() * 100
        ).round(2).reset_index()
        factor_data.columns = [label, "Attrition Rate %"]

        for r_idx, row in enumerate(dataframe_to_rows(factor_data, index=False, header=True), row_start + 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        apply_header_style(ws, row_start + 1, 2)
        row_start += len(factor_data) + 3


def export_salary_sheet(wb: Workbook, df: pd.DataFrame):
    """Create salary analysis sheet."""
    ws = wb.create_sheet("Salary Analysis")

    ws.cell(row=1, column=1, value="Salary Analysis").font = TITLE_FONT
    ws.merge_cells("A1:G1")

    # Salary by department
    dept_salary = df.groupby("department", observed=True).agg(
        Count=("employee_id", "count"),
        Avg_Salary=("monthly_income", "mean"),
        Median_Salary=("monthly_income", "median"),
        Min_Salary=("monthly_income", "min"),
        Max_Salary=("monthly_income", "max"),
        Std_Dev=("monthly_income", "std")
    ).round(2).reset_index()
    dept_salary.columns = ["Department", "Count", "Avg Salary", "Median Salary",
                           "Min Salary", "Max Salary", "Std Dev"]

    ws.cell(row=3, column=1, value="Salary by Department").font = SECTION_FONT
    for r_idx, row in enumerate(dataframe_to_rows(dept_salary, index=False, header=True), 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, 4, len(dept_salary.columns))
    apply_data_styles(ws, 5, 4 + len(dept_salary), len(dept_salary.columns))

    # Gender pay gap
    row_start = 6 + len(dept_salary)
    ws.cell(row=row_start, column=1, value="Gender Pay Gap by Department").font = SECTION_FONT

    gender_pay = df.groupby(["department", "gender"], observed=True).agg(
        Avg_Salary=("monthly_income", "mean"),
        Count=("employee_id", "count")
    ).round(2).reset_index()
    gender_pay.columns = ["Department", "Gender", "Avg Salary", "Count"]

    for r_idx, row in enumerate(dataframe_to_rows(gender_pay, index=False, header=True), row_start + 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row_start + 1, len(gender_pay.columns))


def export_csv_reports(df: pd.DataFrame):
    """Export CSV reports for each major analysis area."""
    print("\nExporting CSV reports...")
    csv_dir = REPORTS_EXCEL_DIR  # Reuse the directory
    os.makedirs(csv_dir, exist_ok=True)

    reports = {
        "employee_data": df,
        "department_summary": df.groupby("department", observed=True).agg(
            employee_count=("employee_id", "count"),
            avg_salary=("monthly_income", "mean"),
            avg_age=("age", "mean"),
            attrition_rate=("attrition", lambda x: (x == "Yes").mean() * 100)
        ).round(2).reset_index(),
        "attrition_data": df[df["attrition"] == "Yes"],
        "active_employees": df[df["attrition"] == "No"],
        "salary_summary": df.groupby("department", observed=True)["monthly_income"].describe().round(2).reset_index()
    }

    for name, data in reports.items():
        filepath = os.path.join(csv_dir, f"{name}.csv")
        data.to_csv(filepath, index=False)
        print(f"  [OK] {filepath}")


def export_summary_report(df: pd.DataFrame):
    """Generate a text-based summary report."""
    kpis = calculate_kpis(df)
    report_lines = [
        "=" * 60,
        "  HR ANALYTICS - EXECUTIVE SUMMARY REPORT",
        "=" * 60,
        f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"  Dataset: {len(df):,} Employees",
        "",
        "-" * 60,
        "  KEY PERFORMANCE INDICATORS",
        "-" * 60,
        f"  Total Employees:        {kpis['total_employees']:,}",
        f"  Active Employees:       {kpis['active_employees']:,}",
        f"  Attrition Rate:         {kpis['attrition_rate']:.1f}%",
        f"  Average Salary:         ${kpis['avg_salary']:,.2f}",
        f"  Average Age:            {kpis['avg_age']:.1f} years",
        f"  Average Tenure:         {kpis['avg_tenure']:.1f} years",
        f"  Avg Job Satisfaction:   {kpis['avg_job_satisfaction']:.2f}/4",
        f"  Avg Work-Life Balance:  {kpis['avg_work_life_balance']:.2f}/4",
        f"  Promotion Rate:         {kpis['promotion_rate']:.1f}%",
        f"  Overtime %:             {kpis['overtime_percentage']:.1f}%",
        f"  Female %:               {kpis['female_percentage']:.1f}%",
        "",
        "-" * 60,
        "  DEPARTMENT HIGHLIGHTS",
        "-" * 60,
    ]

    dept_stats = df.groupby("department").agg(
        count=("employee_id", "count"),
        avg_salary=("monthly_income", "mean"),
        attrition=("attrition", lambda x: (x == "Yes").mean() * 100)
    ).round(2).sort_values("count", ascending=False)

    for dept in dept_stats.index:
        stat = dept_stats.loc[dept]
        report_lines.append(
            f"  {dept:<30s} {int(stat['count']):4d} emp  "
            f"${stat['avg_salary']:>8,.0f}  {stat['attrition']:5.1f}% attrit"
        )

    report_lines.extend([
        "",
        "-" * 60,
        "  TOP INSIGHTS",
        "-" * 60,
        f"  1. Highest attrition department: "
        f"{dept_stats['attrition'].idxmax()} ({dept_stats['attrition'].max():.1f}%)",
        f"  2. Highest paying department: "
        f"{dept_stats['avg_salary'].idxmax()} (${dept_stats['avg_salary'].max():,.0f})",
        f"  3. Overall female representation: {kpis['female_percentage']:.1f}%",
        f"  4. Employees working overtime: {kpis['overtime_percentage']:.1f}%",
        f"  5. Average years since last promotion: {kpis['avg_years_since_promotion']:.1f}",
        "",
        "=" * 60,
        "  END OF REPORT",
        "=" * 60
    ])

    report_text = "\n".join(report_lines)

    # Save as .txt
    txt_path = os.path.join(REPORTS_EXCEL_DIR, "executive_summary.txt")
    with open(txt_path, "w") as f:
        f.write(report_text)
    print(f"\n[SAVED] Summary report: {txt_path}")

    return report_text


def export_all_reports():
    """Main export function - generates all reports."""
    print_section("Report Export")

    # Load data
    cleaned_path = os.path.join(DATA_CLEANED_DIR, "hr_data_cleaned.csv")
    if os.path.exists(cleaned_path):
        df = pd.read_csv(cleaned_path)
        print(f"Loaded cleaned data: {cleaned_path}")
    else:
        df = load_raw_data()
        df = standardize_columns(df)
        print("Loaded raw data and standardized columns")

    df = df.copy()
    kpis = calculate_kpis(df)

    # Export summary
    export_summary_report(df)

    # Export CSV reports
    export_csv_reports(df)

    # Export Excel if available
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        export_kpi_sheet(wb, kpis)
        export_department_sheet(wb, df)
        export_attrition_sheet(wb, df)
        export_salary_sheet(wb, df)

        xlsx_path = os.path.join(REPORTS_EXCEL_DIR, "HR_Analytics_Report.xlsx")
        wb.save(xlsx_path)
        print(f"\n[SAVED] Excel report: {xlsx_path}")
    else:
        print("\n[SKIPPED] Excel report (openpyxl not available)")

    print(f"\nAll reports saved to: {REPORTS_EXCEL_DIR}")


if __name__ == "__main__":
    export_all_reports()
